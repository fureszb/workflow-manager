from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, extract
from datetime import datetime, timedelta
from typing import Optional, List
from io import BytesIO

# Excel export
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# PDF export
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.linecharts import HorizontalLineChart
from reportlab.graphics.charts.legends import Legend

from app.core.database import get_db
from app.models.models import (
    ProcessInstance,
    ProcessType,
    StatusDefinition,
    Email,
    TokenUsage,
    ScriptRun,
)
from app.schemas.schemas import (
    ProcessStatsResponse,
    ProcessStatsByStatus,
    ProcessStatsByMonth,
    EmailStatsResponse,
    EmailStatsByImportance,
    EmailStatsByDay,
    TokenStatsResponse,
    TokenStatsByDay,
    TokenStatsByProvider,
    StatisticsExportRequest,
)

router = APIRouter(prefix="/statistics")


def parse_date_range(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
) -> tuple[datetime, datetime]:
    """Parse date range from query params, default to last 30 days."""
    if end_date:
        end = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
    else:
        end = datetime.now()

    if start_date:
        start = datetime.strptime(start_date, "%Y-%m-%d").replace(hour=0, minute=0, second=0)
    else:
        start = end - timedelta(days=30)

    return start, end


@router.get("/processes", response_model=ProcessStatsResponse)
def process_stats(
    start_date: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    db: Session = Depends(get_db),
):
    """Get process statistics with optional date range filter."""
    start, end = parse_date_range(start_date, end_date)

    # Total processes in range
    base_query = db.query(ProcessInstance).filter(
        ProcessInstance.created_at >= start,
        ProcessInstance.created_at <= end,
    )

    total = base_query.count()

    # Get completed status
    completed_status = db.query(StatusDefinition).filter(
        StatusDefinition.name == "Kész"
    ).first()

    completed = 0
    if completed_status:
        completed = base_query.filter(
            ProcessInstance.status_id == completed_status.id
        ).count()

    in_progress = total - completed

    # By status breakdown
    status_counts = (
        db.query(
            StatusDefinition.name,
            StatusDefinition.color,
            func.count(ProcessInstance.id).label("count"),
        )
        .outerjoin(ProcessInstance, ProcessInstance.status_id == StatusDefinition.id)
        .filter(
            StatusDefinition.is_active == True,
        )
        .group_by(StatusDefinition.id, StatusDefinition.name, StatusDefinition.color)
        .all()
    )

    by_status = [
        ProcessStatsByStatus(name=name, count=count or 0, color=color)
        for name, color, count in status_counts
    ]

    # Monthly trend (last 6 months)
    six_months_ago = datetime.now() - timedelta(days=180)
    monthly_data = (
        db.query(
            ProcessInstance.year,
            ProcessInstance.month,
            func.count(ProcessInstance.id).label("count"),
        )
        .filter(ProcessInstance.created_at >= six_months_ago)
        .group_by(ProcessInstance.year, ProcessInstance.month)
        .order_by(ProcessInstance.year, ProcessInstance.month)
        .all()
    )

    by_month = [
        ProcessStatsByMonth(
            month=f"{year}-{str(month).zfill(2)}",
            count=count,
        )
        for year, month, count in monthly_data
    ]

    # Average completion time (for completed processes)
    avg_completion_days = None
    if completed_status:
        completed_with_times = (
            db.query(ProcessInstance)
            .filter(
                ProcessInstance.status_id == completed_status.id,
                ProcessInstance.started_at.isnot(None),
                ProcessInstance.completed_at.isnot(None),
            )
            .all()
        )

        if completed_with_times:
            total_days = sum(
                (p.completed_at - p.started_at).days
                for p in completed_with_times
            )
            avg_completion_days = round(total_days / len(completed_with_times), 1)

    return ProcessStatsResponse(
        total=total,
        completed=completed,
        in_progress=in_progress,
        by_status=by_status,
        by_month=by_month,
        avg_completion_days=avg_completion_days,
    )


@router.get("/emails", response_model=EmailStatsResponse)
def email_stats(
    start_date: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    db: Session = Depends(get_db),
):
    """Get email statistics with optional date range filter."""
    start, end = parse_date_range(start_date, end_date)

    # Base query with date filter
    base_query = db.query(Email).filter(
        Email.created_at >= start,
        Email.created_at <= end,
    )

    total = base_query.count()
    unread = base_query.filter(Email.is_read == False).count()
    read = total - unread

    # By importance
    importance_counts = (
        db.query(
            Email.importance,
            func.count(Email.id).label("count"),
        )
        .filter(
            Email.created_at >= start,
            Email.created_at <= end,
        )
        .group_by(Email.importance)
        .all()
    )

    # Define importance colors
    importance_colors = {
        "Alacsony": "#6b7280",
        "Közepes": "#3b82f6",
        "Magas": "#f59e0b",
        "Kritikus": "#ef4444",
    }

    by_importance = [
        EmailStatsByImportance(
            importance=importance or "Ismeretlen",
            count=count,
            color=importance_colors.get(importance, "#6b7280"),
        )
        for importance, count in importance_counts
    ]

    # Daily trend (within date range)
    daily_data = (
        db.query(
            func.date(Email.created_at).label("date"),
            func.count(Email.id).label("count"),
        )
        .filter(
            Email.created_at >= start,
            Email.created_at <= end,
        )
        .group_by(func.date(Email.created_at))
        .order_by(func.date(Email.created_at))
        .all()
    )

    by_day = [
        EmailStatsByDay(
            date=str(date),
            count=count,
        )
        for date, count in daily_data
    ]

    # Response rate (if we track responses - using read as proxy)
    response_rate = round((read / total * 100), 1) if total > 0 else 0.0

    return EmailStatsResponse(
        total=total,
        unread=unread,
        read=read,
        by_importance=by_importance,
        by_day=by_day,
        response_rate=response_rate,
    )


@router.get("/tokens", response_model=TokenStatsResponse)
def token_stats(
    start_date: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    db: Session = Depends(get_db),
):
    """Get token usage statistics with optional date range filter."""
    start, end = parse_date_range(start_date, end_date)

    # Base query with date filter
    base_query = db.query(TokenUsage).filter(
        TokenUsage.created_at >= start,
        TokenUsage.created_at <= end,
    )

    # Totals
    totals = db.query(
        func.coalesce(func.sum(TokenUsage.input_tokens), 0).label("input"),
        func.coalesce(func.sum(TokenUsage.output_tokens), 0).label("output"),
        func.coalesce(func.sum(TokenUsage.cost), 0.0).label("cost"),
    ).filter(
        TokenUsage.created_at >= start,
        TokenUsage.created_at <= end,
    ).first()

    total_input = int(totals.input) if totals else 0
    total_output = int(totals.output) if totals else 0
    total_cost = float(totals.cost) if totals else 0.0

    # By provider
    provider_data = (
        db.query(
            TokenUsage.provider,
            func.sum(TokenUsage.input_tokens).label("input_tokens"),
            func.sum(TokenUsage.output_tokens).label("output_tokens"),
            func.sum(TokenUsage.cost).label("cost"),
        )
        .filter(
            TokenUsage.created_at >= start,
            TokenUsage.created_at <= end,
        )
        .group_by(TokenUsage.provider)
        .all()
    )

    # Provider colors
    provider_colors = {
        "ollama": "#10b981",
        "openrouter": "#8b5cf6",
        "openai": "#10a37f",
        "anthropic": "#d97706",
    }

    by_provider = [
        TokenStatsByProvider(
            provider=provider,
            input_tokens=int(input_tokens or 0),
            output_tokens=int(output_tokens or 0),
            cost=float(cost or 0),
            color=provider_colors.get(provider, "#6b7280"),
        )
        for provider, input_tokens, output_tokens, cost in provider_data
    ]

    # Daily trend
    daily_data = (
        db.query(
            func.date(TokenUsage.created_at).label("date"),
            func.sum(TokenUsage.input_tokens).label("input_tokens"),
            func.sum(TokenUsage.output_tokens).label("output_tokens"),
            func.sum(TokenUsage.cost).label("cost"),
        )
        .filter(
            TokenUsage.created_at >= start,
            TokenUsage.created_at <= end,
        )
        .group_by(func.date(TokenUsage.created_at))
        .order_by(func.date(TokenUsage.created_at))
        .all()
    )

    by_day = [
        TokenStatsByDay(
            date=str(date),
            input_tokens=int(input_tokens or 0),
            output_tokens=int(output_tokens or 0),
            cost=float(cost or 0),
        )
        for date, input_tokens, output_tokens, cost in daily_data
    ]

    return TokenStatsResponse(
        total_input_tokens=total_input,
        total_output_tokens=total_output,
        total_cost=total_cost,
        by_provider=by_provider,
        by_day=by_day,
    )


@router.post("/export")
def export_stats(
    request: StatisticsExportRequest,
    db: Session = Depends(get_db),
):
    """Export statistics as PDF or Excel."""
    # Validate format
    if request.format not in ("pdf", "excel"):
        raise HTTPException(status_code=400, detail="Format must be 'pdf' or 'excel'")

    # Get date range
    start, end = parse_date_range(request.start_date, request.end_date)

    # Fetch all statistics data
    process_data = _get_process_stats_data(db, start, end)
    email_data = _get_email_stats_data(db, start, end)
    token_data = _get_token_stats_data(db, start, end)

    if request.format == "excel":
        return _generate_excel_export(process_data, email_data, token_data, start, end)
    else:
        return _generate_pdf_export(process_data, email_data, token_data, start, end)


def _get_process_stats_data(db: Session, start: datetime, end: datetime) -> dict:
    """Get process statistics data for export."""
    base_query = db.query(ProcessInstance).filter(
        ProcessInstance.created_at >= start,
        ProcessInstance.created_at <= end,
    )

    total = base_query.count()

    completed_status = db.query(StatusDefinition).filter(
        StatusDefinition.name == "Kész"
    ).first()

    completed = 0
    if completed_status:
        completed = base_query.filter(
            ProcessInstance.status_id == completed_status.id
        ).count()

    in_progress = total - completed

    # By status breakdown
    status_counts = (
        db.query(
            StatusDefinition.name,
            StatusDefinition.color,
            func.count(ProcessInstance.id).label("count"),
        )
        .outerjoin(ProcessInstance, ProcessInstance.status_id == StatusDefinition.id)
        .filter(StatusDefinition.is_active == True)
        .group_by(StatusDefinition.id, StatusDefinition.name, StatusDefinition.color)
        .all()
    )

    by_status = [
        {"name": name, "count": count or 0, "color": color}
        for name, color, count in status_counts
    ]

    # Monthly trend
    six_months_ago = datetime.now() - timedelta(days=180)
    monthly_data = (
        db.query(
            ProcessInstance.year,
            ProcessInstance.month,
            func.count(ProcessInstance.id).label("count"),
        )
        .filter(ProcessInstance.created_at >= six_months_ago)
        .group_by(ProcessInstance.year, ProcessInstance.month)
        .order_by(ProcessInstance.year, ProcessInstance.month)
        .all()
    )

    by_month = [
        {"month": f"{year}-{str(month).zfill(2)}", "count": count}
        for year, month, count in monthly_data
    ]

    return {
        "total": total,
        "completed": completed,
        "in_progress": in_progress,
        "by_status": by_status,
        "by_month": by_month,
    }


def _get_email_stats_data(db: Session, start: datetime, end: datetime) -> dict:
    """Get email statistics data for export."""
    base_query = db.query(Email).filter(
        Email.created_at >= start,
        Email.created_at <= end,
    )

    total = base_query.count()
    unread = base_query.filter(Email.is_read == False).count()
    read = total - unread

    # By importance
    importance_counts = (
        db.query(
            Email.importance,
            func.count(Email.id).label("count"),
        )
        .filter(
            Email.created_at >= start,
            Email.created_at <= end,
        )
        .group_by(Email.importance)
        .all()
    )

    by_importance = [
        {"importance": importance or "Ismeretlen", "count": count}
        for importance, count in importance_counts
    ]

    # Daily trend
    daily_data = (
        db.query(
            func.date(Email.created_at).label("date"),
            func.count(Email.id).label("count"),
        )
        .filter(
            Email.created_at >= start,
            Email.created_at <= end,
        )
        .group_by(func.date(Email.created_at))
        .order_by(func.date(Email.created_at))
        .all()
    )

    by_day = [
        {"date": str(date), "count": count}
        for date, count in daily_data
    ]

    response_rate = round((read / total * 100), 1) if total > 0 else 0.0

    return {
        "total": total,
        "unread": unread,
        "read": read,
        "by_importance": by_importance,
        "by_day": by_day,
        "response_rate": response_rate,
    }


def _get_token_stats_data(db: Session, start: datetime, end: datetime) -> dict:
    """Get token usage statistics data for export."""
    totals = db.query(
        func.coalesce(func.sum(TokenUsage.input_tokens), 0).label("input"),
        func.coalesce(func.sum(TokenUsage.output_tokens), 0).label("output"),
        func.coalesce(func.sum(TokenUsage.cost), 0.0).label("cost"),
    ).filter(
        TokenUsage.created_at >= start,
        TokenUsage.created_at <= end,
    ).first()

    total_input = int(totals.input) if totals else 0
    total_output = int(totals.output) if totals else 0
    total_cost = float(totals.cost) if totals else 0.0

    # By provider
    provider_data = (
        db.query(
            TokenUsage.provider,
            func.sum(TokenUsage.input_tokens).label("input_tokens"),
            func.sum(TokenUsage.output_tokens).label("output_tokens"),
            func.sum(TokenUsage.cost).label("cost"),
        )
        .filter(
            TokenUsage.created_at >= start,
            TokenUsage.created_at <= end,
        )
        .group_by(TokenUsage.provider)
        .all()
    )

    by_provider = [
        {
            "provider": provider,
            "input_tokens": int(input_tokens or 0),
            "output_tokens": int(output_tokens or 0),
            "cost": float(cost or 0),
        }
        for provider, input_tokens, output_tokens, cost in provider_data
    ]

    # Daily trend
    daily_data = (
        db.query(
            func.date(TokenUsage.created_at).label("date"),
            func.sum(TokenUsage.input_tokens).label("input_tokens"),
            func.sum(TokenUsage.output_tokens).label("output_tokens"),
            func.sum(TokenUsage.cost).label("cost"),
        )
        .filter(
            TokenUsage.created_at >= start,
            TokenUsage.created_at <= end,
        )
        .group_by(func.date(TokenUsage.created_at))
        .order_by(func.date(TokenUsage.created_at))
        .all()
    )

    by_day = [
        {
            "date": str(date),
            "input_tokens": int(input_tokens or 0),
            "output_tokens": int(output_tokens or 0),
            "cost": float(cost or 0),
        }
        for date, input_tokens, output_tokens, cost in daily_data
    ]

    return {
        "total_input_tokens": total_input,
        "total_output_tokens": total_output,
        "total_cost": total_cost,
        "by_provider": by_provider,
        "by_day": by_day,
    }


def _generate_excel_export(
    process_data: dict,
    email_data: dict,
    token_data: dict,
    start: datetime,
    end: datetime,
) -> StreamingResponse:
    """Generate Excel export with multiple worksheets."""
    wb = Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def style_header(ws, row=1, cols=None):
        if cols is None:
            cols = ws.max_column
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

    def auto_width(ws):
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    # === Summary Sheet ===
    ws_summary = wb.active
    ws_summary.title = "Összefoglaló"
    ws_summary.append(["Workflow Manager - Statisztikai Riport"])
    ws_summary.append([f"Időszak: {start.strftime('%Y-%m-%d')} - {end.strftime('%Y-%m-%d')}"])
    ws_summary.append([f"Generálva: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    ws_summary.append([])

    ws_summary.append(["Folyamatok"])
    ws_summary.append(["Összes folyamat:", process_data["total"]])
    ws_summary.append(["Befejezett:", process_data["completed"]])
    ws_summary.append(["Folyamatban:", process_data["in_progress"]])
    ws_summary.append([])

    ws_summary.append(["Emailek"])
    ws_summary.append(["Összes email:", email_data["total"]])
    ws_summary.append(["Olvasott:", email_data["read"]])
    ws_summary.append(["Olvasatlan:", email_data["unread"]])
    ws_summary.append(["Olvasási arány:", f"{email_data['response_rate']}%"])
    ws_summary.append([])

    ws_summary.append(["Token Használat"])
    ws_summary.append(["Input tokenek:", token_data["total_input_tokens"]])
    ws_summary.append(["Output tokenek:", token_data["total_output_tokens"]])
    ws_summary.append(["Összes költség:", f"${token_data['total_cost']:.4f}"])

    # Style title
    ws_summary["A1"].font = Font(bold=True, size=14)
    auto_width(ws_summary)

    # === Processes by Status Sheet ===
    ws_process_status = wb.create_sheet("Folyamatok státusz szerint")
    ws_process_status.append(["Státusz", "Darabszám"])
    style_header(ws_process_status)
    for item in process_data["by_status"]:
        ws_process_status.append([item["name"], item["count"]])
    auto_width(ws_process_status)

    # === Processes by Month Sheet ===
    ws_process_month = wb.create_sheet("Folyamatok havi bontásban")
    ws_process_month.append(["Hónap", "Darabszám"])
    style_header(ws_process_month)
    for item in process_data["by_month"]:
        ws_process_month.append([item["month"], item["count"]])
    auto_width(ws_process_month)

    # === Emails by Importance Sheet ===
    ws_email_importance = wb.create_sheet("Emailek fontosság szerint")
    ws_email_importance.append(["Fontosság", "Darabszám"])
    style_header(ws_email_importance)
    for item in email_data["by_importance"]:
        ws_email_importance.append([item["importance"], item["count"]])
    auto_width(ws_email_importance)

    # === Emails by Day Sheet ===
    ws_email_day = wb.create_sheet("Napi email forgalom")
    ws_email_day.append(["Dátum", "Darabszám"])
    style_header(ws_email_day)
    for item in email_data["by_day"]:
        ws_email_day.append([item["date"], item["count"]])
    auto_width(ws_email_day)

    # === Token Usage by Provider Sheet ===
    ws_token_provider = wb.create_sheet("Token használat provider szerint")
    ws_token_provider.append(["Provider", "Input Tokenek", "Output Tokenek", "Költség ($)"])
    style_header(ws_token_provider, cols=4)
    for item in token_data["by_provider"]:
        ws_token_provider.append([
            item["provider"],
            item["input_tokens"],
            item["output_tokens"],
            round(item["cost"], 4),
        ])
    auto_width(ws_token_provider)

    # === Token Usage by Day Sheet ===
    ws_token_day = wb.create_sheet("Napi token használat")
    ws_token_day.append(["Dátum", "Input Tokenek", "Output Tokenek", "Költség ($)"])
    style_header(ws_token_day, cols=4)
    for item in token_data["by_day"]:
        ws_token_day.append([
            item["date"],
            item["input_tokens"],
            item["output_tokens"],
            round(item["cost"], 4),
        ])
    auto_width(ws_token_day)

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"statistics_report_{start.strftime('%Y%m%d')}_{end.strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _generate_pdf_export(
    process_data: dict,
    email_data: dict,
    token_data: dict,
    start: datetime,
    end: datetime,
) -> StreamingResponse:
    """Generate PDF export with charts and tables."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=2 * cm,
        leftMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Heading1"],
        fontSize=18,
        spaceAfter=30,
        alignment=1,  # Center
    )
    heading_style = ParagraphStyle(
        "CustomHeading",
        parent=styles["Heading2"],
        fontSize=14,
        spaceAfter=12,
        spaceBefore=20,
        textColor=colors.HexColor("#4F46E5"),
    )
    normal_style = styles["Normal"]

    elements = []

    # Title
    elements.append(Paragraph("Workflow Manager - Statisztikai Riport", title_style))
    elements.append(Paragraph(
        f"Időszak: {start.strftime('%Y-%m-%d')} - {end.strftime('%Y-%m-%d')}",
        normal_style,
    ))
    elements.append(Paragraph(
        f"Generálva: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        normal_style,
    ))
    elements.append(Spacer(1, 20))

    # Table styles
    table_style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
        ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F3F4F6")),
        ("TEXTCOLOR", (0, 1), (-1, -1), colors.black),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 9),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D1D5DB")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
    ])

    # === Process Statistics ===
    elements.append(Paragraph("Folyamat Statisztikák", heading_style))

    # Summary table
    process_summary = [
        ["Mutató", "Érték"],
        ["Összes folyamat", str(process_data["total"])],
        ["Befejezett", str(process_data["completed"])],
        ["Folyamatban", str(process_data["in_progress"])],
    ]
    t = Table(process_summary, colWidths=[200, 100])
    t.setStyle(table_style)
    elements.append(t)
    elements.append(Spacer(1, 15))

    # By status table
    if process_data["by_status"]:
        elements.append(Paragraph("Folyamatok státusz szerint", normal_style))
        status_data = [["Státusz", "Darabszám"]]
        for item in process_data["by_status"]:
            status_data.append([item["name"], str(item["count"])])
        t = Table(status_data, colWidths=[200, 100])
        t.setStyle(table_style)
        elements.append(t)
        elements.append(Spacer(1, 15))

    # By month table
    if process_data["by_month"]:
        elements.append(Paragraph("Havi folyamat trend", normal_style))
        month_data = [["Hónap", "Darabszám"]]
        for item in process_data["by_month"]:
            month_data.append([item["month"], str(item["count"])])
        t = Table(month_data, colWidths=[200, 100])
        t.setStyle(table_style)
        elements.append(t)
    elements.append(Spacer(1, 20))

    # === Email Statistics ===
    elements.append(Paragraph("Email Statisztikák", heading_style))

    email_summary = [
        ["Mutató", "Érték"],
        ["Összes email", str(email_data["total"])],
        ["Olvasott", str(email_data["read"])],
        ["Olvasatlan", str(email_data["unread"])],
        ["Olvasási arány", f"{email_data['response_rate']}%"],
    ]
    t = Table(email_summary, colWidths=[200, 100])
    t.setStyle(table_style)
    elements.append(t)
    elements.append(Spacer(1, 15))

    # By importance
    if email_data["by_importance"]:
        elements.append(Paragraph("Emailek fontosság szerint", normal_style))
        importance_data = [["Fontosság", "Darabszám"]]
        for item in email_data["by_importance"]:
            importance_data.append([item["importance"], str(item["count"])])
        t = Table(importance_data, colWidths=[200, 100])
        t.setStyle(table_style)
        elements.append(t)
    elements.append(Spacer(1, 20))

    # === Token Usage Statistics ===
    elements.append(Paragraph("Token Használat Statisztikák", heading_style))

    token_summary = [
        ["Mutató", "Érték"],
        ["Input tokenek", f"{token_data['total_input_tokens']:,}"],
        ["Output tokenek", f"{token_data['total_output_tokens']:,}"],
        ["Összes költség", f"${token_data['total_cost']:.4f}"],
    ]
    t = Table(token_summary, colWidths=[200, 150])
    t.setStyle(table_style)
    elements.append(t)
    elements.append(Spacer(1, 15))

    # By provider
    if token_data["by_provider"]:
        elements.append(Paragraph("Token használat provider szerint", normal_style))
        provider_data = [["Provider", "Input", "Output", "Költség ($)"]]
        for item in token_data["by_provider"]:
            provider_data.append([
                item["provider"],
                f"{item['input_tokens']:,}",
                f"{item['output_tokens']:,}",
                f"{item['cost']:.4f}",
            ])
        t = Table(provider_data, colWidths=[100, 100, 100, 100])
        t.setStyle(table_style)
        elements.append(t)

    # Build PDF
    doc.build(elements)
    buffer.seek(0)

    filename = f"statistics_report_{start.strftime('%Y%m%d')}_{end.strftime('%Y%m%d')}.pdf"

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
