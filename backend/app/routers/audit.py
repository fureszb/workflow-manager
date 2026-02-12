"""Audit log router - tracks all API actions for compliance and debugging."""
from fastapi import APIRouter, Depends, Query, HTTPException, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, desc
from datetime import datetime, timedelta
from typing import Optional, List
from io import BytesIO
import json

# Excel export
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.core.database import get_db
from app.models.models import AuditLog, AppSetting
from app.schemas.schemas import AuditLogResponse, AuditLogCreate

router = APIRouter(prefix="/audit-log")


def parse_date_range(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
) -> tuple[datetime, datetime]:
    """Parse date range from query params, default to last 30 days."""
    if end_date:
        end = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
    else:
        end = datetime.now()

    if start_date:
        start = datetime.strptime(start_date, "%Y-%m-%d").replace(hour=0, minute=0, second=0)
    else:
        start = end - timedelta(days=30)

    return start, end


def create_audit_log(
    db: Session,
    action: str,
    entity_type: Optional[str] = None,
    entity_id: Optional[int] = None,
    details: Optional[str] = None,
):
    """Helper function to create an audit log entry."""
    audit_entry = AuditLog(
        action=action,
        entity_type=entity_type,
        entity_id=entity_id,
        details=details,
    )
    db.add(audit_entry)
    db.commit()
    return audit_entry


@router.get("", response_model=List[AuditLogResponse])
def list_audit_logs(
    start_date: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    action: Optional[str] = Query(None, description="Filter by action type"),
    entity_type: Optional[str] = Query(None, description="Filter by entity type"),
    search: Optional[str] = Query(None, description="Search in details"),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500),
    db: Session = Depends(get_db),
):
    """List audit log entries with filtering and pagination."""
    start, end = parse_date_range(start_date, end_date)

    query = db.query(AuditLog).filter(
        AuditLog.created_at >= start,
        AuditLog.created_at <= end,
    )

    if action:
        query = query.filter(AuditLog.action == action)

    if entity_type:
        query = query.filter(AuditLog.entity_type == entity_type)

    if search:
        query = query.filter(AuditLog.details.ilike(f"%{search}%"))

    # Order by newest first
    query = query.order_by(desc(AuditLog.created_at))

    logs = query.offset(skip).limit(limit).all()
    return logs


@router.get("/count")
def count_audit_logs(
    start_date: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    action: Optional[str] = Query(None, description="Filter by action type"),
    entity_type: Optional[str] = Query(None, description="Filter by entity type"),
    search: Optional[str] = Query(None, description="Search in details"),
    db: Session = Depends(get_db),
):
    """Get count of audit log entries matching filters."""
    start, end = parse_date_range(start_date, end_date)

    query = db.query(func.count(AuditLog.id)).filter(
        AuditLog.created_at >= start,
        AuditLog.created_at <= end,
    )

    if action:
        query = query.filter(AuditLog.action == action)

    if entity_type:
        query = query.filter(AuditLog.entity_type == entity_type)

    if search:
        query = query.filter(AuditLog.details.ilike(f"%{search}%"))

    count = query.scalar()
    return {"count": count}


@router.get("/actions")
def list_action_types(db: Session = Depends(get_db)):
    """Get list of distinct action types for filtering."""
    actions = db.query(AuditLog.action).distinct().order_by(AuditLog.action).all()
    return [a[0] for a in actions if a[0]]


@router.get("/entity-types")
def list_entity_types(db: Session = Depends(get_db)):
    """Get list of distinct entity types for filtering."""
    entity_types = db.query(AuditLog.entity_type).distinct().order_by(AuditLog.entity_type).all()
    return [e[0] for e in entity_types if e[0]]


@router.get("/{log_id}", response_model=AuditLogResponse)
def get_audit_log(log_id: int, db: Session = Depends(get_db)):
    """Get a single audit log entry by ID."""
    log = db.query(AuditLog).filter(AuditLog.id == log_id).first()
    if not log:
        raise HTTPException(status_code=404, detail="Audit log entry not found")
    return log


@router.post("/export")
def export_audit_logs(
    format: str = Query("excel", description="Export format: 'csv' or 'excel'"),
    start_date: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    action: Optional[str] = Query(None, description="Filter by action type"),
    entity_type: Optional[str] = Query(None, description="Filter by entity type"),
    db: Session = Depends(get_db),
):
    """Export audit logs as CSV or Excel."""
    if format not in ("csv", "excel"):
        raise HTTPException(status_code=400, detail="Format must be 'csv' or 'excel'")

    start, end = parse_date_range(start_date, end_date)

    query = db.query(AuditLog).filter(
        AuditLog.created_at >= start,
        AuditLog.created_at <= end,
    )

    if action:
        query = query.filter(AuditLog.action == action)

    if entity_type:
        query = query.filter(AuditLog.entity_type == entity_type)

    logs = query.order_by(desc(AuditLog.created_at)).all()

    if format == "csv":
        return _generate_csv_export(logs, start, end)
    else:
        return _generate_excel_export(logs, start, end)


def _generate_csv_export(logs: List[AuditLog], start: datetime, end: datetime) -> StreamingResponse:
    """Generate CSV export of audit logs."""
    import csv

    buffer = BytesIO()
    # Use UTF-8 BOM for Excel compatibility
    buffer.write(b'\xef\xbb\xbf')

    # Write to buffer using TextIOWrapper
    import io
    text_buffer = io.TextIOWrapper(buffer, encoding='utf-8', newline='')
    writer = csv.writer(text_buffer)

    # Header
    writer.writerow(["ID", "Időpont", "Művelet", "Entitás típus", "Entitás ID", "Részletek"])

    # Data rows
    for log in logs:
        writer.writerow([
            log.id,
            log.created_at.strftime("%Y-%m-%d %H:%M:%S") if log.created_at else "",
            log.action or "",
            log.entity_type or "",
            log.entity_id or "",
            log.details or "",
        ])

    text_buffer.flush()
    text_buffer.detach()
    buffer.seek(0)

    filename = f"audit_log_{start.strftime('%Y%m%d')}_{end.strftime('%Y%m%d')}.csv"

    return StreamingResponse(
        buffer,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _generate_excel_export(logs: List[AuditLog], start: datetime, end: datetime) -> StreamingResponse:
    """Generate Excel export of audit logs."""
    wb = Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def style_header(ws, row=1, cols=None):
        if cols is None:
            cols = ws.max_column
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

    def auto_width(ws):
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[column_letter].width = adjusted_width

    # Main sheet
    ws = wb.active
    ws.title = "Audit Log"

    # Title row
    ws.append(["Workflow Manager - Audit Log Export"])
    ws.append([f"Időszak: {start.strftime('%Y-%m-%d')} - {end.strftime('%Y-%m-%d')}"])
    ws.append([f"Exportálva: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    ws.append([f"Összes bejegyzés: {len(logs)}"])
    ws.append([])

    # Header
    ws.append(["ID", "Időpont", "Művelet", "Entitás típus", "Entitás ID", "Részletek"])
    style_header(ws, row=6, cols=6)

    # Data rows
    for log in logs:
        ws.append([
            log.id,
            log.created_at.strftime("%Y-%m-%d %H:%M:%S") if log.created_at else "",
            log.action or "",
            log.entity_type or "",
            log.entity_id or "",
            log.details or "",
        ])

    # Style title
    ws["A1"].font = Font(bold=True, size=14)
    auto_width(ws)

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"audit_log_{start.strftime('%Y%m%d')}_{end.strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def cleanup_old_audit_logs(db: Session, retention_days: int):
    """Delete audit logs older than retention_days."""
    cutoff_date = datetime.now() - timedelta(days=retention_days)
    deleted_count = db.query(AuditLog).filter(AuditLog.created_at < cutoff_date).delete()
    db.commit()
    return deleted_count
